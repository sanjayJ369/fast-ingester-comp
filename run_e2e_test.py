"""
run_e2e_test.py — End-to-end pipeline test with similarity scoring.

1. Ingest Testing Set (Go fast path)
2. Query all questions from Testing Set Questions.xlsx using Ollama
3. Compare generated answers vs ground truth using:
   - Fuzzy string similarity (SequenceMatcher)
   - Semantic similarity (embedding cosine similarity)
   - Document source accuracy
4. Print detailed report

Usage:
    python run_e2e_test.py              # query only (uses existing indexes)
    python run_e2e_test.py --ingest     # re-ingest then query
"""

import asyncio
import sys
import time
import json
import numpy as np
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

from ingestion.embedder import get_model, _embed_batch
from retrieval.retriever import retrieve_all
from query.answerer import answer_all


# ── Load ground truth from Excel ─────────────────────────────────────────────

def load_ground_truth(xlsx_path: str) -> list[dict]:
    """Load questions, expected answers, and source docs from the Testing Set Excel."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    questions = []
    for row in rows[1:]:  # skip header
        if row[0] is None:
            break
        questions.append({
            "question": str(row[0]),
            "expected_answer": str(row[1]),
            "expected_docs": str(row[2]),
        })
    wb.close()
    return questions


# ── Similarity metrics ───────────────────────────────────────────────────────

def fuzzy_similarity(a: str, b: str) -> float:
    """SequenceMatcher ratio (0-1). Good for exact/near-exact matches."""
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def semantic_similarity(text_a: str, text_b: str) -> float:
    """Cosine similarity using the embedding model."""
    from config import EMBED_DOC_PREFIX
    model = get_model()
    vecs = model.encode(
        [f"{EMBED_DOC_PREFIX}{text_a}", f"{EMBED_DOC_PREFIX}{text_b}"],
        normalize_embeddings=True,
        convert_to_numpy=True,
    ).astype(np.float32)
    return float(np.dot(vecs[0], vecs[1]))


def doc_match_score(generated_doc: str, expected_docs: str) -> float:
    """Check if the generated doc_name matches any of the expected documents."""
    if not generated_doc:
        return 0.0
    gen = generated_doc.lower().strip()
    for expected in expected_docs.lower().split("\n"):
        # Extract just the filename part (before comma/page info)
        expected_name = expected.split(",")[0].strip()
        if not expected_name:
            continue
        # Check substring match (filenames may not be exact)
        if gen in expected_name or expected_name in gen:
            return 1.0
        # Fuzzy match for close names
        if SequenceMatcher(None, gen, expected_name).ratio() > 0.7:
            return 0.8
    return 0.0


# ── Main ─────────────────────────────────────────────────────────────────────

async def run_e2e():
    # Optional: re-ingest
    if "--ingest" in sys.argv:
        from fast_ingest import fast_ingest
        print("=" * 70)
        print("  PHASE 1: INGESTION (Go fast path)")
        print("=" * 70)
        timings = fast_ingest("./Testing Set")
        print(f"\nIngestion timings: {timings}\n")

    # Load ground truth
    gt = load_ground_truth("./Testing Set/Testing Set Questions.xlsx")
    questions = [q["question"] for q in gt]

    print("=" * 70)
    print(f"  PHASE 2: QUERY ({len(questions)} questions via Ollama)")
    print("=" * 70)

    # Warm up embedding model
    print("\n[E2E] Loading embedding model...")
    get_model()

    # Retrieve
    t0 = time.perf_counter()
    print("[E2E] Retrieving chunks for all questions...")
    retrieved = await retrieve_all(questions)
    t_retrieve = time.perf_counter() - t0
    print(f"[E2E] Retrieval done in {t_retrieve:.2f}s\n")

    # Answer with Ollama
    t0 = time.perf_counter()
    print("[E2E] Generating answers with Ollama (mistral)...")
    answers = await answer_all(questions, retrieved)
    t_answer = time.perf_counter() - t0
    print(f"[E2E] Answering done in {t_answer:.2f}s\n")

    # ── Scoring ──────────────────────────────────────────────────────────
    print("=" * 70)
    print("  PHASE 3: SIMILARITY SCORING")
    print("=" * 70)

    results = []
    for i, (q, a) in enumerate(zip(gt, answers)):
        generated = a.answer
        expected = q["expected_answer"]

        fuzzy = fuzzy_similarity(generated, expected)
        semantic = semantic_similarity(generated, expected)
        doc_score = doc_match_score(a.doc_name, q["expected_docs"])

        results.append({
            "idx": i + 1,
            "question": q["question"],
            "expected": expected,
            "generated": generated,
            "doc_name": a.doc_name,
            "confidence": a.confidence,
            "fuzzy_sim": fuzzy,
            "semantic_sim": semantic,
            "doc_match": doc_score,
            "chunks_retrieved": len(retrieved[i]),
        })

    # ── Report ───────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("  DETAILED RESULTS")
    print("=" * 70)

    for r in results:
        print(f"\n{'─'*70}")
        print(f"Q{r['idx']}: {r['question']}")
        print(f"{'─'*70}")
        print(f"  Expected : {r['expected'][:200]}{'...' if len(r['expected']) > 200 else ''}")
        print(f"  Generated: {r['generated'][:200]}{'...' if len(r['generated']) > 200 else ''}")
        print(f"  Doc      : {r['doc_name']} (match: {'YES' if r['doc_match'] >= 0.8 else 'NO'})")
        print(f"  Confidence: {r['confidence']}")
        print(f"  Chunks   : {r['chunks_retrieved']} retrieved")
        print(f"  Fuzzy Sim: {r['fuzzy_sim']:.3f}")
        print(f"  Semantic : {r['semantic_sim']:.3f}")

    # ── Summary table ────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("  SUMMARY")
    print("=" * 70)
    print(f"\n{'Q#':<4} {'Fuzzy':>7} {'Semantic':>9} {'DocMatch':>9} {'Conf':>6}  Question")
    print(f"{'──':<4} {'─────':>7} {'────────':>9} {'────────':>9} {'────':>6}  {'─'*40}")

    for r in results:
        q_short = r["question"][:40] + ("..." if len(r["question"]) > 40 else "")
        print(f"Q{r['idx']:<3} {r['fuzzy_sim']:>7.3f} {r['semantic_sim']:>9.3f} "
              f"{r['doc_match']:>9.1f} {r['confidence']:>6}  {q_short}")

    # Averages
    avg_fuzzy = np.mean([r["fuzzy_sim"] for r in results])
    avg_semantic = np.mean([r["semantic_sim"] for r in results])
    avg_doc = np.mean([r["doc_match"] for r in results])
    high_conf = sum(1 for r in results if r["confidence"] == "high")

    print(f"{'──':<4} {'─────':>7} {'────────':>9} {'────────':>9} {'────':>6}")
    print(f"AVG  {avg_fuzzy:>7.3f} {avg_semantic:>9.3f} {avg_doc:>9.1f}  {high_conf}/{len(results)} high")

    # ── Timing summary ───────────────────────────────────────────────────
    print(f"\n{'─'*70}")
    print(f"  Retrieval : {t_retrieve:.2f}s")
    print(f"  LLM       : {t_answer:.2f}s")
    print(f"  Total     : {t_retrieve + t_answer:.2f}s")
    budget = 30.0
    total = t_retrieve + t_answer
    if total < budget:
        print(f"  WITHIN 30s BUDGET ({budget - total:.1f}s to spare)")
    else:
        print(f"  OVER BUDGET by {total - budget:.1f}s")
    print(f"{'─'*70}")

    # Save results to JSON
    out_path = "./data/e2e_results.json"
    with open(out_path, "w") as f:
        json.dump(results, f, indent=2)
    print(f"\nFull results saved to {out_path}")


if __name__ == "__main__":
    asyncio.run(run_e2e())
